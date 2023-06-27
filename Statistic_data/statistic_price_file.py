# 计算price, bom price, octopart price 在多个sheet 中
# 输入型号，从bom，获取pm，从octopart 获取pr， 输出grade, s_price 是bom的一周内的最高价
# （1） pm = 一周内k_price 平均价
# （2） pr = 各个supplier 的最低价
#  (3) grade = pm/pr -> [10, ++]->A, [5,10]->B, [1-5]->C, [0.5-1.0]->D, [0.1-0.5]->E, [--, 0.1]->F
#  s_price=max_price

from openpyxl import load_workbook
import base64
from WRTools import WaitHelp, PathHelp
from IC_stock import IC_stock_excel_read, IC_Stock_excel_write
import numpy as np
import json
from urllib.request import urlopen
import ssl
import Bom_price.bom_price_info
ssl._create_default_https_context = ssl._create_unverified_context


cate_source_file = PathHelp.get_file_path(super_path='TRenesas_MCU_105H', file_name='Task.xlsx')
result_save_file = cate_source_file

# renesas
octopart_file_arr = ['/Users/liuhe/Desktop/progress/TOnsemi/octopart/4-1.xlsx',
                        '/Users/liuhe/Desktop/progress/TOnsemi/octopart/4-2.xlsx',
                        '/Users/liuhe/Desktop/progress/TOnsemi/octopart/4-3.xlsx',
                        '/Users/liuhe/Desktop/progress/TOnsemi/octopart/4-4.xlsx']
bom_file_arr = ['/Users/liuhe/Desktop/progress/THolt/second/04/bom_price.xlsx',
                '/Users/liuhe/Desktop/progress/THolt/second/11/bom_price.xlsx',
                '/Users/liuhe/Desktop/progress/THolt/second/sz/bom_price.xlsx',
                PathHelp.get_file_path('TOnsemi_DiscontinueP4', 'bom_price.xlsx')]

# onsemi
# octopart_file_arr = ['/Users/liuhe/PycharmProjects/YJCX_AI/THolt/octopart_price.xlsx']
# bom_file_arr = ['/Users/liuhe/Desktop/progress/THolt/second/04/bom_price.xlsx',
#                 '/Users/liuhe/Desktop/progress/THolt/second/11/bom_price.xlsx',
#                 '/Users/liuhe/Desktop/progress/THolt/second/sz/bom_price.xlsx',
#                 PathHelp.get_file_path('THolt', 'bom_price.xlsx')]



# 一次汇总bom 的所有sheets， 二维数组，文件列表+sheet 列表
def get_bom_sheets() -> list:
    result = []
    for file in bom_file_arr:
        # 获取工作簿对象
        wb = load_workbook(filename=file)
        # 获取sheet
        sheets_arr = wb.sheetnames
        result.append(sheets_arr)
    return result


# 从bom获取pm us dollar, supplier invalid 忽略掉， 并返回是否有报价
def get_pm(file_index, sheet_index, rate):
    bom_price = 0
    has_supplier = False
    # 获取工作簿对象
    wb = load_workbook(filename=bom_file_arr[file_index])
    # 获取sheet
    ws = wb.worksheets[sheet_index]
    price_arr = []
    max_price = 0
    if ws.cell(1, 1).value is not None:
        has_supplier = True  # 记录是否有报价
    for i in range(ws.min_row, ws.max_row + 1):
        print(f'是否有报价记录 is: {ws.cell(i, 10).value}')
        valid_supplier = (ws.cell(i, 10).value == "TRUE" or ws.cell(i, 10).value)
        # valid_supplier = False
        # # 以汇总价格的当前的一周内判断为准
        # if ws.cell(i, 7).value is not None:
        #     release_time = ws.cell(i, 7).value
        #     valid_time_arr = ['3天内', '1周内']
        #     if valid_time_arr.__contains__(release_time):
        #         valid_supplier = True
        #     else:
        #         numberDays = WaitHelp.daysPassed(release_time)
        #         if 0 <= numberDays <= 17:
        #             valid_supplier = True
        if not valid_supplier:   # 只需要三天内，一周内掉报价
            break
        price_str = ws.cell(i, 6).value
        if price_str is not None:
            if len(price_str) > 0:
                if '￥' in price_str:
                    price_str = price_str.replace('￥', '')
                    price_str = price_str.replace(',', '')
                    price_float = float(price_str)
                elif '＄' in price_str:
                    price_str = price_str.replace('＄', '')
                    price_str = price_str.replace(',', '')
                    price_float = float(price_str)*rate
                else:
                    price_str = price_str.replace(',', '')
                    price_float = float(price_str)
                if price_float > 0:
                    price_arr.append(price_float)
    if len(price_arr) > 0:
        bom_price = np.mean(price_arr)
        max_price = np.max(price_arr)
    print(f'pm is:{bom_price}')
    result_dic = {"bom_price": bom_price, "has_supplier": has_supplier, "max_price": max_price}
    return result_dic


# 一次汇总octopart 的所有sheets， 二维数组，文件列表+sheet 列表
def get_octopart_sheets() -> list:
    result = []
    for file in octopart_file_arr:
        # 获取工作簿对象
        wb = load_workbook(filename=file)
        # 获取sheet
        sheets_arr = wb.sheetnames
        result.append(sheets_arr)
    return result


# 从octopart获取pr, rmb, 把sheetName 和cate 关联, 未授权的supplier，提供的价格主动忽略
def get_pr(file_index, sheet_index, rate) -> float:
    result = 0
    # 获取工作簿对象
    wb = load_workbook(filename=octopart_file_arr[file_index])
    # 获取sheet
    ws = wb.worksheets[sheet_index]
    price_arr = []
    for i in range(ws.min_row, ws.max_row + 1):
        authore_status = ws.cell(i, 9).value
        if authore_status == -1:  # 未授权supplier
            continue
        price_str = ws.cell(i, 9).value
        if price_str and price_str != 'None' and len(price_str) > 0:
            price_str = price_str.replace(',', '')
            price_float = float(price_str) * rate
            if price_float > 0:
                price_arr.append(price_float)
    if len(price_arr) > 0:
        result = np.min(price_arr)
    print(f'pr is:{result}')
    return result


def get_grade(c_value: float) -> str:
    result = ''
    if c_value == 0:
        result = '??'
    elif c_value >= 10:
        result = 'A'
    elif 5 <= c_value < 10:
        result = 'B'
    elif 1 <= c_value < 5:
        result = 'C'
    elif 0.5 <= c_value < 1.0:
        result = 'D'
    elif 0.1 <= c_value < 0.5:
        result = 'E'
    else:
        result = 'F'
    print(f'grade is:{result}')
    return result


# 计算汇率
def get_rate():
    result = 6.94  # default cate
    try:
        url = "https://api.exchangerate-api.com/v4/latest/USD"
        json_str = ''
        resp = urlopen(url)
        resp = resp.read().decode(resp.headers.get_content_charset() or 'ascii')
        json_dic = json.loads(resp)
        result = json_dic['rates']['CNY']
        print(f'net rate is:{result}')
    except Exception as e:
        print('get_rate exception: {e}')
    return result


# 通过sheets 的二维数组，获取文件index， sheet index
def get_indexs(source_arr, cate_name) -> list:
    sheet_name_base64str = str(base64.b64encode(cate_name.encode('utf-8')), 'utf-8')
    for (file_index, file_sheets) in enumerate(source_arr):
        for (sheet_index, sheet_name) in enumerate(file_sheets):
            if sheet_name == sheet_name_base64str:
                return [file_index, sheet_index]
    return None


# 获取pm,pr
# 获取grade 保存
def calculater_price():
    rate = get_rate()
    all_cates = IC_stock_excel_read.get_cate_name_arr(cate_source_file, 'ppn', 1)
    sub_cates = all_cates[0:]
    all_bom_sheets = get_bom_sheets()
    all_octopart_sheets = get_octopart_sheets()
    for (cate_index, cate_name) in enumerate(sub_cates):
        if cate_name is None:
            continue
        bom_index = get_indexs(all_bom_sheets, str(cate_name))
        oct_index = get_indexs(all_octopart_sheets, str(cate_name))
        print(f'cate is:{cate_index} -> {cate_name}, bom_index is:{bom_index} , octopart_index is : {oct_index}')
        if bom_index is not None:
            pm_dic = get_pm(file_index=bom_index[0], sheet_index=bom_index[1], rate=rate)
        else:
            pm_dic = {"bom_price": 0, "has_supplier": False, 'max_price': 0}
        pm = pm_dic['bom_price']
        bom_has_supplier = pm_dic['has_supplier']
        bom_max_price = pm_dic['max_price']
        if oct_index is not None:
            pr = get_pr(file_index=oct_index[0], sheet_index=oct_index[1], rate=rate)
        else:
            pr = 0
        if pm == 0 or pr == 0:
            c = 0
        else:
            c = pm/pr
        grade = get_grade(c)
        manu_name = IC_stock_excel_read.get_cell_content(file_name=cate_source_file, sheet_name='ppn', row=cate_index + 1, col=2)
        row_arr = [[cate_name, manu_name, pm, pr, c, grade, '有' if bom_has_supplier else '无', str(bom_max_price)]]  # 从bom，获取pm，从octopart 获取pr
        IC_Stock_excel_write.add_arr_to_sheet(file_name=result_save_file, sheet_name='bom_octopart_price', dim_arr=row_arr)


# 更新信的octopart——price 和bom-price ，获取新的grade， 新的grade 为A or ？？ 不变，否则删除
def update_gradeA():
    wb = load_workbook(filename=cate_source_file)
    new_ws = wb['new']
    need_delete_cates = []
    # 根据单元格名称获取单元格对象
    for i in range(new_ws.min_row, new_ws.max_row + 1):
        cate = new_ws.cell(i, 1).value
        grade = new_ws.cell(i, 6).value
        if not(grade == "A" or grade == '??'):
            need_delete_cates.append(cate)
    print("need_delete_cates is:", need_delete_cates)
    need_update_sheets = ['before0910']
    for temp_sheet_name in need_update_sheets:
        source_cates = IC_stock_excel_read.get_cate_name_arr(file_name=cate_source_file, sheet_name=temp_sheet_name, col_index=1)
        result = list(set(source_cates).difference(set(need_delete_cates)))
        wb.remove(wb[temp_sheet_name])
        wb.save(filename=cate_source_file)
        wb.create_sheet(temp_sheet_name)
        wb.save(filename=cate_source_file)
        temp_sheet = wb[temp_sheet_name]
        for temp_cate in result:
            temp_sheet.append([temp_cate])
    wb.save(filename=cate_source_file)
    wb.close()


if __name__ == "__main__":
    calculater_price()
    # update_gradeA()
    print('over')


