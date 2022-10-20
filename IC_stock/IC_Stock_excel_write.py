from openpyxl import workbook, load_workbook, Workbook
import os
import base64
from IC_stock import IC_stock_excel_read

default_exel_name = 'E01_stock_IC.xlsx'
if os.path.exists(default_exel_name):
    wb = load_workbook(default_exel_name)
else:
    wb = Workbook()


def create_sheet(sheet_name):
    print(f'add sheet : {sheet_name}')
    wb.create_sheet(sheet_name)


def isSheetExist(sheet_name):
    old_sheet_names = wb.sheetnames
    if old_sheet_names.__contains__(sheet_name):
        return True
    else:
        return False


# 将库存数据保存到sheet 中
# sheet_name: 由cate_name base64加密得到
# arr 是二维数组[[cate1_name, cate1_supplier], [cate2_name, cate2_supplier]]
def add_arr_to_sheet(file_name, sheet_name: object, dim_arr: object):
    global wb
    print('arr is:', dim_arr)
    if os.path.exists(file_name):
        wb = load_workbook(file_name)
    else:
        wb = Workbook()
    if not isSheetExist(sheet_name):
        create_sheet(sheet_name)
    sheet = wb[sheet_name]
    for ele in dim_arr:
        sheet.append(ele)
    wb.save(file_name)
    wb.close()


def set_col_width(sheet_name):
    sheet = wb[sheet_name]
    sheet.column_dimensions["A"].width = 36.0


# 在单元格中写入内容
def write_cell(file_name, sheet_name, row, col, value):
    global wb
    if os.path.exists(file_name):
        wb = load_workbook(file_name)
    else:
        wb = Workbook()
    sheet = wb[sheet_name]
    sheet.cell(row=row, column=col).value = value
    wb.save(file_name)
    wb.close()


# 写入自己创建的sheet 即可避免sheets 为空的问题
def active_excel(file_name, sheet_name):
    global wb
    wb = load_workbook(file_name)
    wb.create_sheet(sheet_name)
    wb.save(file_name)
    wb.close()


# 将Excel 除第一个sheet 外的所哟sheet 删除
def delet_excel(file_name):
    global wb
    wb = load_workbook(file_name)
    for (index, sheet) in enumerate(wb.worksheets):
        if index == 0:
            continue
        else:
            wb.remove(sheet)
    wb.save(file_name)
    wb.close()


# 将cate——sheet 合并到一个sheet 中
def move_sheet_to_row(source_file_list, aim_file, aim_sheet, cate_name):
    for source_file in source_file_list:
        source_wb = load_workbook(source_file)
        source_sheet_names = source_wb.sheetnames
        for (sheet_index, sheet_name) in enumerate(source_sheet_names):
            sheet_name_base64str = str(base64.b64encode(cate_name.encode('utf-8')), 'utf-8')
            if sheet_name_base64str == sheet_name:
                source_list = IC_stock_excel_read.get_sheet_content_by_name(file_name=source_file, sheet_name=sheet_name)
                add_arr_to_sheet(file_name=aim_file, sheet_name=aim_sheet, dim_arr=source_list)
                return


if __name__ == "__main__":
    add_arr_to_sheet(file_name='aa.xlsx', sheet_name='a', dim_arr=[[1, 2]])

