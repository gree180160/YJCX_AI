# 计算manu对应的supplier的排名

from openpyxl import workbook, load_workbook, Workbook
from WRTools import ExcelHelp
import base64
import IC_Stock_Info
from Supplier_analyth import Manu_supp
import time


cate_source_file = '//T1203IC_Stock.xlsx'
result_save_file = cate_source_file
ICStock_file_arr = ['/Users/liuhe/Desktop/progress/T1203IC_stock/T1203_ICStock_1k.xlsx',
                    '/Users/liuhe/Desktop/progress/T1203IC_stock/T1203_ICStock_2k.xlsx',
                    '/Users/liuhe/Desktop/progress/T1203IC_stock/T1203_ICStock_3k.xlsx',
                    '/Users/liuhe/Desktop/progress/T1203IC_stock/T1203_ICStock_4k.xlsx',
                    '/Users/liuhe/Desktop/progress/T1203IC_stock/unfinished/T1203_ICStock_1k.xlsx',
                    '/Users/liuhe/Desktop/progress/T1203IC_stock/unfinished/T1203_ICStock_2k.xlsx',
                    '/Users/liuhe/Desktop/progress/T1203IC_stock/unfinished/T1203_ICStock_3k.xlsx',
                    '/Users/liuhe/Desktop/progress/T1203IC_stock/unfinished/T1203_ICStock_4k.xlsx'
                    ]


# 根据file_index, sheet_index 获取cate 在IC 中的记录, [cate_name, manu, valid_supplier, valid_stock]
def get_cate_stock(file_index, sheet_index, cate_name, manu) -> list:
    # 获取工作簿对象
    file_name = ICStock_file_arr[file_index]
    wb = load_workbook(filename=file_name)
    sheet = wb.worksheets[sheet_index]
    min_row = sheet.min_row
    max_row = sheet.max_row
    # 遍历sheet
    supplier_arr = []
    for row in range(min_row, max_row + 1):
        supplier = sheet.cell(row, 1).value
        if not (supplier is None or supplier == "--"):
            iccp_str = sheet.cell(row, 2).value
            isICCP = "notICCP" not in iccp_str
            sscp_str = sheet.cell(row, 3).value
            isSSCP = "notSSCP" not in sscp_str
            model = sheet.cell(row, 4).value
            isSpotRanking = "notSpotRanking" not in sheet.cell(row, 5).value
            isHotSell = "notHotSell" not in sheet.cell(row, 6).value
            manufacturer = sheet.cell(row, 7).value
            stock_num = sheet.cell(row, 8).value
            search_date = sheet.cell(row, 9).value
            ic_Stock_Info = IC_Stock_Info.IC_Stock_Info(supplier=supplier, isICCP=isICCP, isSSCP=isSSCP,
                                                        model=model,
                                                        isSpotRanking=isSpotRanking, isHotSell=isHotSell,
                                                        manufacturer=manufacturer, stock_num=stock_num,
                                                        search_date=search_date)
            if ic_Stock_Info.is_valid_supplier():
                supplier_arr.append(supplier)
    print(f'cate is: {cate_name} manu is: {manu} supplier is: {supplier_arr}')
    return supplier_arr


# 获取ICStock_file_arr 中的所有sheet
def get_ICStock_sheets():
    result = []
    for file in ICStock_file_arr:
        # 获取工作簿对象
        wb = load_workbook(filename=file)
        # 获取sheet
        sheets_arr = wb.sheetnames
        result.append(sheets_arr)
    return result


# 通过sheets 的二维数组，获取文件index， sheet index
def get_indexs(source_arr, cate_name) -> list:
    sheet_name_base64str = str(base64.b64encode(cate_name.encode('utf-8')), 'utf-8')
    for (file_index, file_sheets) in enumerate(source_arr):
        for (sheet_index, sheet_name) in enumerate(file_sheets):
            if sheet_name == sheet_name_base64str:
                print(f'file_index is:{file_index} sheet_index is:{sheet_index}')
                return [file_index, sheet_index]
    print(f'not found {cate_name}')
    return None


#  统计manu 对应的supplier 排名
def statistic_supplier():
    all_cates = ExcelHelp.read_col_content(file_name=cate_source_file, sheet_name='all_pn', col_index=1)
    sub_cates = all_cates[0:]
    all_ic_sheets = get_ICStock_sheets()
    manu_supp_list = []
    rank_manu_list = []  # [[supplier_name, ppn_count]]
    for (cate_index, cate_name) in enumerate(sub_cates):
        if cate_name is None:
            continue
        print(f'index is: {cate_index}, pn is :{cate_name}')
        ic_index_arr = get_indexs(all_ic_sheets, str(cate_name))
        manu_name = ExcelHelp.read_cell_content(file_name=cate_source_file, sheet_name='all_pn',
                                                         row=cate_index + 1, col=2)

        manu_record_index = None
        for (index, temp_manu_supp) in enumerate(manu_supp_list):
            if temp_manu_supp.manu_name == manu_name:
                manu_record_index = index
                break
        if manu_record_index is not None:
            current_manu_supplier = manu_supp_list[manu_record_index]
        else:
            current_manu_supplier = Manu_supp(manu_name=manu_name)
            manu_supp_list.append(current_manu_supplier)
        if ic_index_arr is not None:
            supplier_name_list = get_cate_stock(file_index=ic_index_arr[0],
                                                sheet_index=ic_index_arr[1],
                                                cate_name=str(cate_name),
                                                manu=manu_name)
        else:
            supplier_name_list = []
        for temp_supplier in supplier_name_list:
            current_manu_supplier.input_record(manu_name=manu_name, supplier_name=temp_supplier)
            rank_manu_list = add_supplierToRank(supplier=temp_supplier, old_rank_list=rank_manu_list)
    for temp_manu in manu_supp_list:
        row_arr = temp_manu.output_record()
        ExcelHelp.add_arr_to_sheet(file_name=result_save_file, sheet_name='manu_supplier', dim_arr=row_arr)
    ExcelHelp.add_arr_to_sheet(file_name=result_save_file, sheet_name='supplie_rank', dim_arr=rank_manu_list)


# 在排名中加入supplier，将重新排名的结果返回
def add_supplierToRank(supplier, old_rank_list) -> list:
    print(f"rank add supplier:{supplier}")
    result = old_rank_list
    has_record = False
    for temp_rank_ele in result:
        if temp_rank_ele[0] == supplier:
            temp_rank_ele[1] += 1
            has_record = True
            break;
    if not has_record:
        temp_rank_ele = [supplier, 1]
        result.append(temp_rank_ele)
    return result


# supplier 出现的总次数排名
def supplier_rank():
    supplier_repeat_list = ExcelHelp.read_col_content(file_name=cate_source_file, sheet_name='manu_supplier', col_index=2)
    supplier_dif_list = list(set(supplier_repeat_list))
    result = []
    for (index_set, supplier_name_set) in enumerate(supplier_dif_list):
        ppn_count = 0
        for (index_repeat, supplier_name_repeat) in enumerate(supplier_repeat_list):
            if supplier_name_repeat == supplier_name_set:
                ppn_count += int(ExcelHelp.read_cell_content(file_name=cate_source_file, sheet_name='manu_supplier', row=index_repeat + 1, col=3))
        supplier_info_list = [supplier_name_set, ppn_count]
        result.append(supplier_info_list)
    ExcelHelp.add_arr_to_sheet(file_name=cate_source_file, sheet_name='supplie_rank', dim_arr=result)


if __name__ == "__main__":
    statistic_supplier()
    # supplier_rank()



