import string
import os
from openpyxl import workbook, load_workbook, Workbook
import base64


# 返回cate list
def get_cate_name_arr(file_name: object, sheet_name: object, col_index: int) -> list:
    # 获取工作簿对象
    wb = load_workbook(filename=file_name)
    # 获取sheet
    ws = wb[sheet_name]
    # 根据单元格名称获取单元格对象
    result = []
    for i in range(ws.min_row, ws.max_row + 1):
        cate = ws.cell(i, col_index).value
        if not cate == "--":
            result.append(cate)
    # print("result is:", result)
    wb.save(filename=file_name)
    wb.close()
    return result


def get_cell_content(file_name, sheet_name, col, row):
    # 获取工作簿对象
    wb = load_workbook(filename=file_name)
    # 获取sheet
    ws = wb[sheet_name]
    # 根据单元格名称获取单元格对象
    cell_content = ws.cell(row=row, column=col).value
    return cell_content


# 获取sheet 中到内容，返回二维数组
def get_sheet_content_by_name(file_name, sheet_name):
    wb = load_workbook(file_name)
    ws = wb[sheet_name]
    # 获取所有行所有列
    content_list = []
    for row in ws.iter_rows():
        row_list = []
        for cell in row:
            cell_value = cell.value
            row_list.append(cell_value)
        if len(row_list) > 0:
            content_list.append(row_list)
    return content_list

# 获取sheet 中到内容，返回二维数组
def get_sheet_content_by_index(file_name, sheet_index):
    wb = load_workbook(file_name)
    ws = wb.worksheets[sheet_index]
    # 获取所有行所有列
    content_list = []
    for row in ws.iter_rows():
        row_list = []
        for cell in row:
            cell_value = cell.value
            row_list.append(cell_value)
        if len(row_list) > 0:
            content_list.append(row_list)
    return content_list


# temp 功能，将ic_stock 结果/2 改成/6
def recalculator(file_name, sheet_name, col):
    # 获取工作簿对象
    wb = load_workbook(filename=file_name)
    # 获取sheet
    ws = wb[sheet_name]
    # 根据单元格名称获取单元格对象
    for i in range(ws.min_row, ws.max_row + 1):
        oldValue = ws.cell(row=i, column=col).value
        if oldValue:
            newValue = int(float(oldValue)*6)
            ws.cell(row=i, column=col).value = str(newValue)
    wb.save(filename=file_name)
    wb.close()


if __name__ == "__main__":
    get_cate_name_arr('T0815.xlsx', 'left', 1)

