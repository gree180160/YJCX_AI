# 计算price
# 输入型号，从bom，获取pm，从octopart 获取pr， 输出grade
# （1） pm = 一周内k_price 平均价
# （2） pr = 各个supplier 的最低价
#  (3) grade = pm/pr -> [10, ++]->A, [5,10]->B, [1-5]->C, [0.5-1.0]->D, [0.1-0.5]->E, [--, 0.1]->F

from openpyxl import load_workbook
import base64
from WRTools import ExcelHelp
from IC_stock import IC_stock_excel_read, IC_Stock_excel_write
import numpy as np
import json
from urllib.request import urlopen
import ssl
ssl._create_default_https_context = ssl._create_unverified_context

cate_source_file = '/Users/liuhe/PycharmProjects/SeleniumDemo/T1012war.xlsx'
result_save_file = cate_source_file
bom_file_arr = ['/Users/liuhe/Desktop/progress/1012war/bom_price_cate_1012war.xlsx']
octopart_file_arr = ['/Users/liuhe/Desktop/progress/1012war/octopart_price_cate_1012war.xlsx']


# 一次汇总bom 的所有sheets， 二维数组，文件列表+sheet 列表
def get_bom_sheets() -> list:
    result = []
    for file in bom_file_arr:
        # 获取工作簿对象
        wb = load_workbook(filename=file)
        # 获取sheet
        sheets_arr = wb.sheetnames
        result.append(sheets_arr)
    return result


# 从bom获取pm us dollar
def get_pm(file_index, sheet_index, rate) -> float:
    result = 0
    # 获取工作簿对象
    wb = load_workbook(filename=bom_file_arr[file_index])
    # 获取sheet
    ws = wb.worksheets[sheet_index]
    price_arr = []
    for i in range(ws.min_row, ws.max_row + 1):
        price_str = ws.cell(i, 6).value
        if price_str is not None:
            if len(price_str) > 0:
                if '￥' in price_str:
                    price_str = price_str.replace('￥', '')
                    price_str = price_str.replace(',', '')
                    price_float = float(price_str)
                elif '＄' in price_str:
                    price_str = price_str.replace('＄', '')
                    price_str = price_str.replace(',', '')
                    price_float = float(price_str)*rate
                if price_float > 0:
                    price_arr.append(price_float)
    if len(price_arr) > 0:
        result = np.mean(price_arr)
    # print(f'pm is:{result}')
    return result


# 一次汇总octopart 的所有sheets， 二维数组，文件列表+sheet 列表
def get_octopart_sheets() -> list:
    result = []
    for file in octopart_file_arr:
        # 获取工作簿对象
        wb = load_workbook(filename=file)
        # 获取sheet
        sheets_arr = wb.sheetnames
        result.append(sheets_arr)
    return result


# 从octopart获取pr, rmb, 把sheetName 和cate 关联
def get_pr(file_index, sheet_index, rate) -> float:
    result = 0
    # 获取工作簿对象
    wb = load_workbook(filename=octopart_file_arr[file_index])
    # 获取sheet
    ws = wb.worksheets[sheet_index]
    price_arr = []
    for i in range(ws.min_row, ws.max_row + 1):
        price_str = ws.cell(i, 9).value
        if price_str is not None and len(price_str) > 0:
            price_str = price_str.replace(',', '')
            price_float = float(price_str) * rate
            if price_float > 0:
                price_arr.append(price_float)
    if len(price_arr) > 0:
        result = np.min(price_arr)
    print(f'pr is:{result}')
    return result


def get_grade(c_value: float) -> str:
    result = ''
    if c_value == 0:
        result = '??'
    elif c_value >= 10:
        result = 'A'
    elif 5 <= c_value < 10:
        result = 'B'
    elif 1 <= c_value < 5:
        result = 'C'
    elif 0.5 <= c_value < 1.0:
        result = 'D'
    elif 0.1 <= c_value < 0.5:
        result = 'E'
    else:
        result = 'F'
    print(f'grade is:{result}')
    return result


# 计算汇率
def get_rate():
    result = 6.85  # default cate
    try:
        url = "https://api.exchangerate-api.com/v4/latest/USD"
        json_str = ''
        resp = urlopen(url)
        resp = resp.read().decode(resp.headers.get_content_charset() or 'ascii')
        json_dic = json.loads(resp)
        result = json_dic['rates']['CNY']
        print(f'net rate is:{result}')
    except Exception as e:
        print('get_rate exception: {e}')
    return result


# 通过sheets 的二维数组，获取文件index， sheet index
def get_indexs(source_arr, cate_name) -> list:
    sheet_name_base64str = str(base64.b64encode(cate_name.encode('utf-8')), 'utf-8')
    for (file_index, file_sheets) in enumerate(source_arr):
        for (sheet_index, sheet_name) in enumerate(file_sheets):
            if sheet_name == sheet_name_base64str:
                return [file_index, sheet_index]
    return None


# 获取pm,pr
# 获取grade 保存
def calculater_price():
    rate = get_rate()
    all_cates = IC_stock_excel_read.get_cate_name_arr(cate_source_file, 'all', 1)
    sub_cates = all_cates[0:]
    all_bom_sheets = get_bom_sheets()
    all_octopart_sheets = get_octopart_sheets()
    for (cate_index, cate_name) in enumerate(all_cates):
        if cate_name is None:
            continue
        bom_index = get_indexs(all_bom_sheets, str(cate_name))
        oct_index = get_indexs(all_octopart_sheets, str(cate_name))
        print(f'cate is:{cate_index} -> {cate_name}, bom_index is:{bom_index} , octopart_index is : {oct_index}')
        if bom_index is not None:
            pm = get_pm(file_index=bom_index[0], sheet_index=bom_index[1], rate = rate)
        else:
            pm = 0
        if oct_index is not None:
            pr = get_pr(file_index=oct_index[0], sheet_index=oct_index[1], rate=rate)
        else:
            pr = 0
        if pm == 0 or pr == 0:
            c = 0
        else:
            c = pm/pr
        grade = get_grade(c)
        manu_name = IC_stock_excel_read.get_cell_content(file_name=cate_source_file, sheet_name='all', row=cate_index + 1, col=2)
        row_arr = [[cate_name, manu_name, pm, pr, c, grade]] #从bom，获取pm，从octopart 获取pr
        IC_Stock_excel_write.add_arr_to_sheet(file_name=result_save_file, sheet_name='price', dim_arr=row_arr)


# 更新信的octopart——price 和bom-price ，获取新的grade， 新的grade 为A or ？？ 不变，否则删除
def update_gradeA():
    wb = load_workbook(filename=cate_source_file)
    new_ws = wb['new']
    need_delete_cates = []
    # 根据单元格名称获取单元格对象
    for i in range(new_ws.min_row, new_ws.max_row + 1):
        cate = new_ws.cell(i, 1).value
        grade = new_ws.cell(i, 6).value
        if not(grade == "A" or grade == '??'):
            need_delete_cates.append(cate)
    print("need_delete_cates is:", need_delete_cates)
    need_update_sheets = ['before0910']
    for temp_sheet_name in need_update_sheets:
        source_cates = IC_stock_excel_read.get_cate_name_arr(file_name=cate_source_file, sheet_name=temp_sheet_name, col_index=1)
        result = list(set(source_cates).difference(set(need_delete_cates)))
        wb.remove(wb[temp_sheet_name])
        wb.save(filename=cate_source_file)
        wb.create_sheet(temp_sheet_name)
        wb.save(filename=cate_source_file)
        temp_sheet = wb[temp_sheet_name]
        for temp_cate in result:
            temp_sheet.append([temp_cate])
    wb.save(filename=cate_source_file)
    wb.close()



if __name__ == "__main__":
    calculater_price()
    # update_gradeA()
    print('over')


