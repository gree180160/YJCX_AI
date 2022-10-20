#  记录E01 提供的型号，在IC 中的库存信息
import base64
from selenium import webdriver
from selenium.webdriver.common.by import By
import time
import random
import undetected_chromedriver as uc
import ssl
from IC_stock import IC_stock_excel_read, IC_Stock_excel_write
from WRTools import LogHelper, UserAgentHelper, IPHelper, WaitHelp
from arrow_transition_info import Arrow_transition_info
from openpyxl import load_workbook

ssl._create_default_https_context = ssl._create_unverified_context

total_page = 1
current_page = 1
accouts_arr = [["helen@molies.net", "Molies741258!"]]
driver_option = webdriver.ChromeOptions()
# driver_option.add_argument(f'--proxy-server=http://{WRTools.IPHelper.getRandowIP()}')
driver_option.add_argument("–incognito")
#  等待初始HTML文档完全加载和解析，
driver_option.page_load_strategy = 'eager'
# driver_option.add_argument(f'user-agent="{WRTools.UserAgentHelper.getRandowUA()}"')
prefs = {"profile.managed_default_content_settings.images": 2}
driver_option.add_experimental_option('prefs', prefs)
driver = uc.Chrome(use_subprocess=True)
# driver.set_page_load_timeout(10000)
default_url = 'https://my.arrow.com/en-US/products/search/PCM5100APWR?q=PCM5100APWR&limit=10&page=1'
cate_source_file = '/Users/liuhe/PycharmProjects/SeleniumDemo/T_gradeA_json.xlsx'
result_save_file = '/Users/liuhe/PycharmProjects/SeleniumDemo/Arrow_transition/arrow_transition_cate_08.xlsx'
log_file = '/Users/liuhe/PycharmProjects/SeleniumDemo/Arrow_transition/arrow_transition_log.txt'


# arrow page
def get_total_page():
    global total_page
    global current_page
    total_page = 1
    try:
        li_arr = driver.find_element(by=By.CLASS_NAME,
                                     value="pagination__summary-title.ng-untouched ng-pristine.ng-valid").find_elements(
            by=By.TAG_NAME, value='span')
        if total_page is not None and len(li_arr) > 0:
            total_page = int(li_arr[len(li_arr) - 1].text)
    except:
        li_arr = []
        total_page = 1
    current_page = 1


# myarrow_login
def login_action(aim_url):
    current_url = driver.current_url
    if current_url == "https://my.arrow.com/en-US/login":
        # begin login
        accout_current = random.choice(accouts_arr)
        driver.find_element(by=By.ID, value='form-field-1').clear()
        driver.find_element(by=By.ID, value='form-field-1').send_keys(accout_current[0])
        WaitHelp.waitfor_account_import(False, False)
        driver.find_element(by=By.ID, value='form-field-2').clear()
        driver.find_element(by=By.ID, value='form-field-2').send_keys(accout_current[1])
        WaitHelp.waitfor_account_import(False, False)
        driver.find_element(by=By.CLASS_NAME, value='btn.btn-primary').click()
        WaitHelp.waitfor_account_import(True, False)
    elif current_url == 'https://my.arrow.com/en-US/not-found':
        driver.get("https://my.arrow.com/en-US/login")
        WaitHelp.waitfor_account_import(False, False)
        login_action(aim_url)
    # 登陆完成，强行跳转到cate——detail页面
    if not driver.current_url.startswith('https://my.arrow.com/en-US/products/search'):
        driver.get(aim_url)
    WaitHelp.waitfor_account_import(True, False)


# # myarrow account important
# def waitfor(is_load_page):
#     # load new page
#     if is_load_page:
#         time.sleep(150 + random.randint(1, 120))
#     else:
#         time.sleep(5 + random.randint(1, 10))


# 跳转到下一个指定的型号
def go_to_cate(cate_index, cate_name):
    try:
        # 在搜索过程中
        if driver.current_url.startswith('https://my.arrow.com/en-US/products/search/'):
            if cate_name in driver.current_url:
                return
            form = driver.find_element(by=By.CSS_SELECTOR, value='form')  #  view first form is search bar
            input_area = form.find_element(By.CSS_SELECTOR, value='input')
            input_area.clear()
            input_area.send_keys(cate_name)
            search_button = form.find_element(By.CSS_SELECTOR, value='button')
            search_button.click()
            # 延时几秒确保页面加载完毕
            WaitHelp.waitfor_account_import(True, False)
        else:
            driver.get(f"https://my.arrow.com/en-US/products/search/{cate_name}?q={cate_name}&limit=50&page=1")
            # 延时几秒确保页面加载完毕
            WaitHelp.waitfor_account_import(True, False)
    except Exception as e:
        LogHelper.write_log(log_file_name=log_file, content=f'{cate_name} go_to_cate except: {e}')


#   二维数组，page 列表， table 列表
def get_stock(cate_index, cate_name):
    global current_page
    login_action(f"https://my.arrow.com/en-US/products/search/{cate_name}?q={cate_name}&limit=50&page={current_page}")
    get_total_page()
    print(f"cate_index is: {cate_index} cate is:{cate_name} currentPage is: {current_page} totalpage is:{total_page}")
    #  2-loop page arr
    need_load_nextPage = True
    while current_page <= total_page and need_load_nextPage:
        try:
            table = driver.find_elements(by=By.TAG_NAME, value='table')[1]
            tr_list = table.find_elements(by=By.TAG_NAME, value='tr')
        except:
            tr_list = []
        need_save_ic_arr = []
        #  3-loop table arr
        for tempTr in tr_list:
            # print("li is:", templi.__str__())
            td_list = tempTr.find_elements(by=By.TAG_NAME, value='td')
            if len(td_list) >= 8:
                # cate, manufacture, ManuPartNum, public, buffer, multiple, spq, pipeline, pip_quantity, lead_time
                try:
                    manufacture = td_list[2].text
                except:
                    manufacture = '--'
                try:
                    ManuPartNum = td_list[1].find_element(by=By.CSS_SELECTOR, value='div.product-row__description').text
                except:
                    ManuPartNum = '--'
                try:
                    app_inventory = td_list[6].find_element(by=By.TAG_NAME, value='app-inventory')
                    div_list = app_inventory.find_elements(by=By.TAG_NAME, value='div')
                    div_list_count = len(div_list)
                    if div_list_count == 0:
                        public = buffer = multiple = spq = '--'
                    elif div_list_count < 4:
                        public = div_list[0 % div_list_count].text
                        buffer = div_list[1 % div_list_count].text
                        multiple = div_list[2 % div_list_count].text
                        spq = div_list[3 % div_list_count].text
                    else:
                        public = div_list[0].text
                        buffer = div_list[1].text
                        multiple = div_list[2].text
                        spq = div_list[3].text
                except:
                    public = buffer = multiple = spq = '--'
                try:
                    pipeline_all_info = td_list[7].find_element(by=By.TAG_NAME, value='div').text
                except:
                    pipeline_all_info = '--'
                try:
                    lead_time = td_list[8].find_element(by=By.TAG_NAME, value='div').text
                except:
                    lead_time = '--'
                arr_transition_obj = Arrow_transition_info(cate=cate_name, manufacture=manufacture,
                                                           ManuPartNum=ManuPartNum, public_str=public,
                                                           buffer_str=buffer, multiple_str=multiple, spq_str=spq,
                                                           pipeline_all_info=pipeline_all_info, lead_time=lead_time)
                saveContent_arr = arr_transition_obj.descritpion_arr()
                need_save_ic_arr.append(saveContent_arr)
        # save per page
        IC_Stock_excel_write.add_arr_to_sheet(file_name=result_save_file, sheet_name="myarrow",
                                              dim_arr=need_save_ic_arr)
        need_save_ic_arr.clear()
        # 翻页
        old_url = driver.current_url
        if current_page < total_page:
            try:
                page_div = driver.find_element(by=By.CSS_SELECTOR, value='div.pagination__navigation')
                nextButton = page_div.find_elements(by=By.TAG_NAME, value='button')[-1]
                nextButton.click()
            except:
                login_action(
                    f"https://my.arrow.com/en-US/products/search/{cate_name}?q={cate_name}&limit=50&page={current_page}")
            WaitHelp.waitfor_account_import(True, False)
            waitTime = 0  # wait reload time
            while driver.current_url == old_url and waitTime <= 3:
                WaitHelp.waitfor_account_import(False, False)
                waitTime += 1
                print("long page:", driver.current_url)
        current_page += 1


# 获取garde == A 的所有cate_name
def get_A_cate(file_name_arr) -> list:
    result = []
    for file_name in file_name_arr:
        # 获取工作簿对象
        wb = load_workbook(filename=file_name)
        # 获取sheet
        ws = wb['price']
        # 根据单元格名称获取单元格对象
        for i in range(ws.min_row, ws.max_row + 1):
            cate_name = ws.cell(i, 1).value
            grade = ws.cell(i, 6).value
            if grade == "A":
                result.append(cate_name)
    wb.close()
    return result


def main():
    # all_cates = get_A_cate(cate_source_file_list)
    all_cates = IC_stock_excel_read.get_cate_name_arr(file_name=cate_source_file,sheet_name='06and09', col_index=1)
    cate_ids = all_cates[0:]
    for (cate_index, cate_name) in enumerate(cate_ids):
        if cate_name is None:
            continue
        print(f'cate_index is: {cate_index}  cate_name is: {cate_name}')
        if cate_index > 0 and cate_index%15 == 0:
            time.sleep(500)
        go_to_cate(cate_index, cate_name)
        get_stock(cate_index, cate_name)


if __name__ == "__main__":
    driver.get(default_url)
    WaitHelp.waitfor_account_import(True, False)
    # 先登陆
    login_action(default_url)
    # 开始搜索
    main()