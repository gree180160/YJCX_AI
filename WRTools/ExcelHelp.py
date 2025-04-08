from openpyxl import load_workbook, Workbook
import openpyxl
import os
import time
from datetime import datetime, timedelta
from openpyxl.styles import PatternFill

# READ
# 获取某一列的内容返回cate list
def read_col_content(file_name: str, sheet_name: str, col_index: int) -> list:
    # 获取工作簿对象
    wb = load_workbook(filename=file_name)
    # 获取sheet
    ws = wb[sheet_name]
    # 根据单元格名称获取单元格对象
    result = []
    for i in range(ws.min_row, ws.max_row + 1):
        cate = ws.cell(i, col_index).value
        if not cate == "--":
            result.append(str(cate))
    # print("result is:", result)
    wb.save(filename=file_name)
    wb.close()
    return result


# READ
# 获取某一行的内容返回cate list
def read_row_content(file_name: str, sheet_name: str, row_index: int) -> list:
    # 获取工作簿对象
    wb = load_workbook(filename=file_name)
    # 获取sheet
    ws = wb[sheet_name]
    # 根据单元格名称获取单元格对象
    result = []
    rows = ws[row_index]
    for r in rows:
        result.append(str(r.value))
    wb.save(filename=file_name)
    wb.close()
    return result


# 获取某一cell的内容 返回str
def read_cell_content(file_name, sheet_name, col, row) -> str:
    # 获取工作簿对象
    wb = load_workbook(filename=file_name)
    # 获取sheet
    ws = wb[sheet_name]
    # 根据单元格名称获取单元格对象
    cell_content = ws.cell(row=row, column=col).value
    return str(cell_content)


# 根据sheet_name 获取sheet 中到内容，返回二维数组
def read_sheet_content_by_name(file_name, sheet_name):
    wb = load_workbook(file_name)
    ws = wb[sheet_name]
    # 获取所有行所有列
    content_list = []
    for row in ws.iter_rows():
        row_list = []
        for cell in row:
            cell_value = str(cell.value)
            row_list.append(cell_value)
        if len(row_list) > 0:
            content_list.append(row_list)
    return content_list


# 读取sheet
#  eg:从第3列到第5列，from_col=3， to_col = 5,
def read_from_col_to_col(file_name, sheet_name, from_col, to_col):
    wb = load_workbook(file_name)
    ws = wb[sheet_name]
    # 获取所有行所有列
    content_list = []
    for row in ws.iter_rows():
        row_list = []
        for (index, cell) in enumerate(row):
            if to_col > index >= from_col - 1:
                cell_value = str(cell.value)
                row_list.append(cell_value)
        if len(row_list) > 0:
            content_list.append(row_list)
    return content_list


# WRITE
def create_excel_file(file_path):
    workbook = openpyxl.Workbook()
    workbook.save(file_path)


def create_sheet(sheet_name, wb):
    print(f'add sheet : {sheet_name}')
    wb.create_sheet(sheet_name)


def isSheetExist(sheet_name, wb):
    old_sheet_names = wb.sheetnames
    if old_sheet_names.__contains__(sheet_name):
        return True
    else:
        return False


# 将库存数据保存到sheet 中
# sheet_name
# arr 是二维数组[[cate1_name, cate1_supplier], [cate2_name, cate2_supplier]]
def add_arr_to_sheet(file_name, sheet_name: object, dim_arr: object):
    print('arr is:', dim_arr)
    if os.path.exists(file_name):
        wb = load_workbook(file_name)
    else:
        wb = Workbook()
    if not isSheetExist(sheet_name, wb):
        create_sheet(sheet_name, wb)
    sheet = wb[sheet_name]
    for ele in dim_arr:
        sheet.append(ele)
    wb.save(file_name)
    wb.close()


# 将库存数据保存到sheet 中
# sheet_name
# arr 是一维数组[cate1_name, cate2_name]
def add_arr_to_col(file_name, sheet_name: object, dim_arr: object):
    print('arr is:', dim_arr)
    if os.path.exists(file_name):
        wb = load_workbook(file_name)
    else:
        wb = Workbook()
    if not isSheetExist(sheet_name, wb):
        create_sheet(sheet_name, wb)
    sheet = wb[sheet_name]
    for ele in dim_arr:
        sheet.append([ele])
    wb.save(file_name)
    wb.close()


# 把一个数组保存到某一列中
# 从第二行开始写入，skip_row_count = 1
def save_one_col(file_name: str, sheet_name: str, col_index: int, dim_arr: list, skip_row_count:int):
    print('arr is:', dim_arr)
    if os.path.exists(file_name):
        wb = load_workbook(file_name)
    else:
        wb = Workbook()
    if not isSheetExist(sheet_name, wb):
        create_sheet(sheet_name, wb)
    sheet = wb[sheet_name]
    for (ele_index, ele_contents) in enumerate(dim_arr):
        for (value_index, value) in enumerate(ele_contents):
            try:
                if value and len(value) > 0:
                    sheet.cell(row=ele_index + 1 + skip_row_count, column=col_index+value_index).value = value
            except:
                print(f'{value_index}th ele: {value} set is err')
    wb.save(file_name)
    wb.close()


# 把一个数组保存到sheet 中，从指定到列开始保存
# dim_arr：二维数组
def save_content_from_col(file_name: str, sheet_name: str, start_col: int, dim_arr: list):
    print('arr is:', dim_arr)
    if os.path.exists(file_name):
        wb = load_workbook(file_name)
    else:
        wb = Workbook()
    if not isSheetExist(sheet_name, wb):
        create_sheet(sheet_name, wb)
    sheet = wb[sheet_name]
    for (ele_index, row_values) in enumerate(dim_arr):
        try:
            if row_values and len(row_values) > 0:
                for (index, cell_value) in enumerate(row_values):
                    sheet.cell(row=ele_index + 1, column=start_col + index).value = cell_value
        except:
            print(f'{ele_index}th ele: {row_values} set is err')
    wb.save(file_name)
    wb.close()


def set_col_width(sheet_name, wb):
    sheet = wb[sheet_name]
    sheet.column_dimensions["A"].width = 36.0


# 在单元格中写入内容
def write_cell(file_name, sheet_name, row, col, value):
    if os.path.exists(file_name):
        wb = load_workbook(file_name)
    else:
        wb = Workbook()
    sheet = wb[sheet_name]
    sheet.cell(row=row, column=col).value = value
    wb.save(file_name)
    wb.close()


# 写入自己创建的sheet 即可避免sheets 为空的问题
def active_excel(file_name, sheet_name):
    # wb = load_workbook(file_name)
    # wb.create_sheet(sheet_name)
    # wb.save(file_name)
    # wb.close()
    arr = [["active test"]]
    add_arr_to_sheet(file_name=file_name, sheet_name=sheet_name, dim_arr=arr)


def delete_sheet_content(file_name, sheet_name):
    wb = load_workbook(file_name)
    for (index, sheet) in enumerate(wb.worksheets):
        if sheet.title == sheet_name:
            sheet = wb[sheet_name]
            sheet.delete_rows(1, sheet.max_row)
    wb.save(file_name)
    wb.close()



def remove_sheet(file_name, sheet_name):
    wb = load_workbook(file_name)
    for (index, sheet) in enumerate(wb.worksheets):
        if sheet.title == sheet_name:
            wb.remove(sheet)
    wb.save(file_name)
    wb.close()


# 将Excel 除第一个sheet 外的所哟sheet 删除
def remove_sheets(file_name):
    wb = load_workbook(file_name)
    for (index, sheet) in enumerate(wb.worksheets):
        if index == 0:
            continue
        else:
            wb.remove(sheet)
    wb.save(file_name)
    wb.close()


# 读取某个sheet 内容，然后写到另一个sheet 中，可以用于多个sheet 合并到同一个sheet 中
def move_sheet_to_row(source_file_list, source_sheet, aim_file, aim_sheet):
    for source_file in source_file_list:
        source_wb = load_workbook(source_file)
        source_sheet_names = source_wb.sheetnames
        for (sheet_index, sheet_name) in enumerate(source_sheet_names):
            sheet_name_base64str = source_sheet
            if sheet_name_base64str == sheet_name:
                source_list = read_sheet_content_by_name(file_name=source_file, sheet_name=sheet_name)
                add_arr_to_sheet(file_name=aim_file, sheet_name=aim_sheet, dim_arr=source_list)
                return


# octopart search pns 过滤正则出来的关键字的查找结果
# 1.长度大于50，则放弃
# 全部数字，则放弃
# X-Y, 则只保留Y
def deal_keyword_result():
    source = read_col_content(file_name='keyresults.xlsx', sheet_name='all', col_index=1)
    pn_arr = []
    for tempPN in source:
        if str(tempPN).__len__() > 50 or str(tempPN).__len__() < 3:
            continue
        else:
            if str(tempPN).isdigit():
                continue
            else:
                sub_arr = str(tempPN).split('-')
                if len(sub_arr) > 1:
                    pn_value = sub_arr[-1]
                    pn_arr.append([pn_value])
                else:
                    pn_arr.append([tempPN])
    add_arr_to_sheet(file_name='keyresults.xlsx', sheet_name='pns', dim_arr=pn_arr)


# IC——search
def get_search_num(file_name, sheet_name, col_index):
    # 获取工作簿对象
    wb = load_workbook(filename=file_name)
    # 获取sheet
    ws = wb[sheet_name]
    # 根据单元格名称获取单元格对象
    result = []
    for i in range(ws.min_row + 1, ws.max_row + 1):
        cate = ws.cell(i, col_index).value
        if not cate == "--" and not cate == "/":
            result.append(cate)
    # print("result is:", result)
    wb.save(filename=file_name)
    wb.close()
    return result


# 判断sheet 是否为空
def sheet_isEmpty(file_name, sheet_name1, sheet_name2):
    # 获取工作簿对象
    wb = load_workbook(filename=file_name)
    # 获取sheet
    ws1 = wb[sheet_name1]
    # 根据单元格名称获取单元格对象
    print(f'ws1 value {ws1.cell(1, 1).value}')

    # 获取sheet
    ws2 = wb[sheet_name2]
    # 根据单元格名称获取单元格对象
    print(f'ws2 value {ws2.cell(1, 1).value}')


# 将excel 中的所有sheet 内容合并到一个sheet 中
def mergeSheet(source_file, aim_sheet):
        temp_wb = load_workbook(filename=source_file)
        # 获取sheet
        temp_sheets_arr = temp_wb.sheetnames
        for temp_sheet in temp_sheets_arr:
            if temp_sheet != aim_sheet:
                sheet_content = read_sheet_content_by_name(file_name=source_file, sheet_name=temp_sheet)
                add_arr_to_sheet(file_name=source_file, sheet_name=aim_sheet, dim_arr=sheet_content)


def mergeExcel(source_files, aim_file):
    for temp_file in source_files:
        # 获取工作簿对象
        temp_wb = load_workbook(filename=temp_file)
        # 获取sheet
        temp_sheets_arr = temp_wb.sheetnames
        for temp_sheet in temp_sheets_arr:
            sheet_content = read_sheet_content_by_name(file_name=temp_file, sheet_name=temp_sheet)
            add_arr_to_sheet(file_name=aim_file, sheet_name=temp_sheet, dim_arr=sheet_content)


def set_col_width(file_name, sheet_name, width):
    workbook = openpyxl.load_workbook(file_name)
    sheet = workbook[sheet_name]
    for col in sheet.columns:
        column_letter = col[0].column_letter  # 获取列的字母标识（如 'A', 'B' 等）
        sheet.column_dimensions[column_letter].width = width
    workbook.save(file_name)
    workbook.close()


def set_col_color(file_name, sheet_name, color_str, startCol, endCol):
    workbook = openpyxl.load_workbook(file_name)
    # 选择 Sheet1
    sheet = workbook[sheet_name]
    # 定义颜色
    blue_fill = PatternFill(start_color=color_str, end_color=color_str, fill_type='solid')
    # 设置第 1、2 列背景色为蓝色
    for row in sheet.iter_rows(min_col=startCol, max_col=endCol):  # 遍历第 1、2 列
        for cell in row:
            cell.fill = blue_fill
    # 保存修改后的文件
    workbook.save('your_file_modified.xlsx')
    workbook.close()


def render_date():
    files = ['/Users/liuhe/Downloads/tender_info_2023-07-27_B.xlsx', '/Users/liuhe/Downloads/tender_info_2023-07-27_A.xlsx']
    for temp_file in files:
        sheetContent = read_sheet_content_by_name(file_name=temp_file, sheet_name='tender')
        new_content = []
        for row_content in sheetContent:
            new_row_value = row_content
            new_row_value[3] = row_content[3].replace(' ', '')
            new_row_value[4] = row_content[4].replace(' ', '')
            new_row_value[5] = row_content[5].replace(' ', '')
            new_content.append(new_row_value)
        add_arr_to_sheet(file_name=temp_file, sheet_name='Sheet', dim_arr=new_content)

if __name__ == "__main__":
    source_file = '/Users/liuhe/PycharmProjects/YJCX_AI/TICHot.xlsx'
    sheet_content = read_sheet_content_by_name(source_file, 'jdmc')
    print(sheet_content[0])
    print(sheet_content[1])
