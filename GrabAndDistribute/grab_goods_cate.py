import Arrow_transition
import Findchips_stock
from IC_stock import IC_stock_excel_read, IC_Stock_excel_write
from openpyxl import load_workbook


# 获取garde == A 的所有cate_name
def get_A_cate(file_name_arr) -> list:
    result = []
    for file_name in file_name_arr:
        # 获取工作簿对象
        wb = load_workbook(filename=file_name)
        # 获取sheet
        ws = wb['price']
        # 根据单元格名称获取单元格对象
        for i in range(ws.min_row, ws.max_row + 1):
            cate_name = ws.cell(i, 1).value
            grade = ws.cell(i, 6).value
            if grade == "A":
                result.append(cate_name)
    wb.close()
    return result


def get_need_cates(a_source_fils: list, manu_file, finish_file):
    cate_source_file_list = a_source_fils
    a_cates = get_A_cate(cate_source_file_list)
    json_cates = IC_stock_excel_read.get_cate_name_arr(file_name=manu_file, sheet_name='json', col_index=1)
    finished_cateds = IC_stock_excel_read.get_cate_name_arr(file_name=finish_file, sheet_name='all', col_index=1)
    all_cates = list(set(a_cates).union(set(json_cates)))
    all_cates = list(set(all_cates).difference(set(finished_cateds)))
    result = []
    for cate_name in all_cates:
        result.append([cate_name])
    # print(f'findchips result is: {result}')
    return result


def get_findchips_left():
    finished_file = '/Users/liuhe/PycharmProjects/SeleniumDemo/T0806.xlsx'
    cate_source_file_list = ['/Users/liuhe/PycharmProjects/SeleniumDemo/T0829zmz.xlsx',
                             '/Users/liuhe/PycharmProjects/SeleniumDemo/T0815.xlsx']
    json_file = '/Users/liuhe/PycharmProjects/SeleniumDemo/T0906.xlsx'
    findchips_cates = get_need_cates(a_source_fils=cate_source_file_list, manu_file=json_file,
                                     finish_file=finished_file)
    IC_Stock_excel_write.add_arr_to_sheet(file_name=json_file, sheet_name="FindchipsLeft",
                                          dim_arr=findchips_cates)


def get_arrow_left():
    json_file = '/Users/liuhe/PycharmProjects/SeleniumDemo/T0906.xlsx'
    cate_source_file_list = ['/Users/liuhe/PycharmProjects/SeleniumDemo/T0829zmz.xlsx',
                             '/Users/liuhe/PycharmProjects/SeleniumDemo/T0815.xlsx',
                             '/Users/liuhe/PycharmProjects/SeleniumDemo/T0806.xlsx']
    json_cates = IC_stock_excel_read.get_cate_name_arr(file_name=json_file, sheet_name='json', col_index=1)
    a_cates = get_A_cate(cate_source_file_list)
    # print(f'finished_cates count is:{len(a_cates)}')
    all_cates = list(set(json_cates).union(set(a_cates)))
    result = []
    for cate_name in all_cates:
        result.append([cate_name])
    IC_Stock_excel_write.add_arr_to_sheet(file_name=json_file, sheet_name="ArrowLeft",
                                          dim_arr=result)


# 将0808/15/29的结果中cate——garde 为A的结果和findchips left 的结果汇总到一个sheet
def combine_findchips_08():
    T0906_cates_file = '/Users/liuhe/PycharmProjects/SeleniumDemo/T0906.xlsx'
    aim_cates = IC_stock_excel_read.get_cate_name_arr(file_name=T0906_cates_file, sheet_name='ArrowLeft', col_index=1)
    # 1.将0806的有用结果汇总到 T0906_cates_file
    for (cate_index, cate_name) in enumerate(aim_cates):
        if cate_name is None:
            continue
        IC_Stock_excel_write.move_sheet_to_row(source_file_list=[
            '/Users/liuhe/PycharmProjects/SeleniumDemo/Findchips_stock/findchips_stock_cate_0806.xlsx',
            '/Users/liuhe/PycharmProjects/SeleniumDemo/Findchips_stock/findchips_stock_cate_08062.xlsx',
            '/Users/liuhe/PycharmProjects/SeleniumDemo/Findchips_stock/findchips_stock_cate_0906.xlsx',
            '/Users/liuhe/PycharmProjects/SeleniumDemo/Findchips_stock/findchips_stock_cate_09062.xlsx'],
                                               aim_file='/Users/liuhe/PycharmProjects/SeleniumDemo/Grab_goods/grab_goods_0906.xlsx',
                                               aim_sheet='findchips',
                                               cate_name=cate_name)
    print('08 over')


def combine_findchips_0906():
    T0906_cates_files = ['/Users/liuhe/PycharmProjects/SeleniumDemo/Findchips_stock/findchips_stock_cate_0906.xlsx',
                         '/Users/liuhe/PycharmProjects/SeleniumDemo/Findchips_stock/findchips_stock_cate_09062.xlsx']
    result_file = '/Users/liuhe/PycharmProjects/SeleniumDemo/Grab_goods/grab_goods_0906.xlsx'
    T0906_cates_file = '/Users/liuhe/PycharmProjects/SeleniumDemo/T0906.xlsx'
    aim_cates = IC_stock_excel_read.get_cate_name_arr(file_name=T0906_cates_file, sheet_name='ArrowLeft', col_index=1)
    for file in T0906_cates_files:
        wb = load_workbook(file)
        for (index, sheet_name) in enumerate(wb.sheetnames):
            # if sheet_name
            sheetContent_list = IC_stock_excel_read.get_sheet_content_by_index(file_name=file, sheet_index=index)
            IC_Stock_excel_write.add_arr_to_sheet(file_name=result_file, sheet_name="findchips",
                                                  dim_arr=sheetContent_list)
    print('0906 over')


def get0907left():
    source_file = '/Users/liuhe/PycharmProjects/SeleniumDemo/T0907zmz.xlsx'
    finish_files = ['/Users/liuhe/PycharmProjects/SeleniumDemo/T0829zmz.xlsx',
                    '/Users/liuhe/PycharmProjects/SeleniumDemo/T0815.xlsx',
                    '/Users/liuhe/PycharmProjects/SeleniumDemo/T0806.xlsx']
    T29 = IC_stock_excel_read.get_cate_name_arr(file_name=finish_files[0], sheet_name='all', col_index=1)
    T15 = IC_stock_excel_read.get_cate_name_arr(file_name=finish_files[0], sheet_name='all', col_index=1)
    T06 = IC_stock_excel_read.get_cate_name_arr(file_name=finish_files[0], sheet_name='all', col_index=1)
    finsish_cates_set = set(T29).union(set(T15)).union(T06)
    source_cates = IC_stock_excel_read.get_cate_name_arr(file_name=source_file, sheet_name='all', col_index=1)
    all_cates = list(set(source_cates).difference(finsish_cates_set))
    result = []
    for cate_name in all_cates:
        index = source_cates.index(cate_name)
        cate_manu = IC_stock_excel_read.get_cell_content(file_name=source_file, sheet_name='all', col=2, row=index + 1)
        result.append([cate_name, cate_manu])
    IC_Stock_excel_write.add_arr_to_sheet(file_name=source_file, sheet_name="task",
                                          dim_arr=result)


def where_error():
    error_cate = 'ACS712ELCTR-20A-T'
    error_md5 = 'QUNTNzEyRUxDVFItMjBBLVQ='
    file1 = '/Users/liuhe/PycharmProjects/SeleniumDemo/Findchips_stock/findchips_stock_cate_0906.xlsx'
    wb1 = load_workbook(file1)
    sheets1 = wb1.sheetnames
    if error_md5 in sheets1:
        print('error in 1')
    file2 = '/Users/liuhe/PycharmProjects/SeleniumDemo/Findchips_stock/findchips_stock_cate_09062.xlsx'
    wb2 = load_workbook(file2)
    sheets2 = wb2.sheetnames
    if error_md5 in sheets2:
        print('error in 2')


#  查找grade 为A的cate， 然后保存到指定文件
def saveA(save_file, save_sheet):
    all_A_cates = get_A_cate(['/Users/liuhe/PycharmProjects/SeleniumDemo/T0806.xlsx',
                              '/Users/liuhe/PycharmProjects/SeleniumDemo/T0815.xlsx',
                              '/Users/liuhe/PycharmProjects/SeleniumDemo/T0829zmz.xlsx',
                              '/Users/liuhe/PycharmProjects/SeleniumDemo/T0907zmz.xlsx',
                              '/Users/liuhe/PycharmProjects/SeleniumDemo/T0909.xlsx'])
    result = []
    for cate_name in all_A_cates:
        result.append([cate_name])
    IC_Stock_excel_write.add_arr_to_sheet(file_name=save_file, sheet_name=save_sheet,
                                          dim_arr=result)


# 从型号表price 中gradeA的cate ，union TJson_recommand 中的cates
def get_jsonAndGradeA_cates() -> list:
    task_cates_files = ['/Users/liuhe/PycharmProjects/SeleniumDemo/T0806.xlsx',
                        '/Users/liuhe/PycharmProjects/SeleniumDemo/T0815.xlsx',
                        '/Users/liuhe/PycharmProjects/SeleniumDemo/T0829zmz.xlsx',
                        '/Users/liuhe/PycharmProjects/SeleniumDemo/T0907zmz.xlsx',
                        '/Users/liuhe/PycharmProjects/SeleniumDemo/T0909.xlsx']
    gradeA_cates = get_A_cate(file_name_arr=task_cates_files)
    jsonRecommnad_file = '/Users/liuhe/PycharmProjects/SeleniumDemo/TJson_recommand.xlsx'
    jsonRecommand_cates = IC_stock_excel_read.get_cate_name_arr(file_name=jsonRecommnad_file, sheet_name='all',
                                                                col_index=1)
    result = list(set(gradeA_cates).union(set(jsonRecommand_cates)))
    return result


# 更新后的gradeA 汇总名单
def new_jsonAndGradeA_cates() -> list:
    gradeA_file = '/Users/liuhe/PycharmProjects/SeleniumDemo/TgardeA.xlsx'
    json_file = '/Users/liuhe/PycharmProjects/SeleniumDemo/TJson_recommand.xlsx'
    new_gradeA_cates = IC_stock_excel_read.get_cate_name_arr(file_name=gradeA_file, sheet_name='all',
                                                             col_index=1)
    jsonRecommand_cates = IC_stock_excel_read.get_cate_name_arr(file_name=json_file, sheet_name='all', col_index=1)
    black_cates = IC_stock_excel_read.get_cate_name_arr(file_name=json_file, sheet_name='blacklist', col_index=1)
    result_red = list(set(new_gradeA_cates).union(set(jsonRecommand_cates)))
    result = list(set(result_red).difference(set(black_cates)))
    return result


if __name__ == '__main__':
    #  get_findchips_left()
    # combine_findchips_08()
    # where_error()
    # get0907left()
    saveA(save_file='/Users/liuhe/PycharmProjects/SeleniumDemo/TgardeA.xlsx', save_sheet='sum')
