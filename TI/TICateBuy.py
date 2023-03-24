from selenium import webdriver
from selenium.webdriver.common.by import By
import time
import undetected_chromedriver as uc
import ssl
import re
from WRTools import IPHelper, UserAgentHelper, WaitHelp, EmailHelper, LogHelper
from IC_stock import IC_stock_excel_read, IC_Stock_excel_write
from openpyxl import load_workbook
import json

cate_source_file = '//TJson_recommand.xlsx'
result_save_file = '//TI/TI_cate_add.xlsx'
gradeA_file_arr = ['/Users/liuhe/PycharmProjects/YJCX_AI/T0806.xlsx',
                   '/Users/liuhe/PycharmProjects/YJCX_AI/T0815.xlsx',
                   '/Users/liuhe/PycharmProjects/YJCX_AI/T0829zmz.xlsx',
                   '/Users/liuhe/PycharmProjects/YJCX_AI/T0907zmz.xlsx']
log_file = '//TI/TI_buy_log.txt'

ssl._create_default_https_context = ssl._create_unverified_context
option = webdriver.ChromeOptions()
driver = uc.Chrome(use_subprocess=True)
default_url = 'https://www.ti.com.cn/sitesearch/zh-cn/docs/universalsearch.tsp?langPref=zh-CN&searchTerm=TCAN1043DEVM&nr=16#q=TCAN1043DEVM&numberOfResults=25'


# 查询结果显示在which element 中， {1:ti-opn-snapshot, 2:ti-tool-snapshot, 3:ti-gpn-snapshot}
def use_snapshot_kind() -> int:
    result = None
    search_snap = driver.find_element(by=By.CSS_SELECTOR, value='div.search-query-snapshots')
    opn_snap = search_snap.find_element(By.CSS_SELECTOR, value='ti-opn-snapshot')
    if opn_snap.is_displayed():
        result = 1
    else:
        tool_snap = search_snap.find_element(By.CSS_SELECTOR, value='ti-tool-snapshot')
        if tool_snap.is_displayed():
            result = 2
        else:
            gpn_snap = search_snap.find_element(By.CSS_SELECTOR, value='ti-gpn-snapshot')
            if gpn_snap.is_displayed():
                result = 3
    return result

# 判断是否需要登陆，需要就点击登陆按钮
def login_action(aim_url):
    if driver.current_url.__contains__('https://www.ti.com/sitesearch') or driver.current_url.__contains__('https://www.ti.com.cn/sitesearch'):
        try:
            ti_login = driver.find_element(by=By.CSS_SELECTOR, value='ti-login')
            login_shadow = ti_login.shadow_root
            a = login_shadow.find_element(by=By.CSS_SELECTOR, value='a.ti-login-link')
            webdriver.ActionChains(driver).move_to_element(a).click(a).perform()
            WaitHelp.waitfor_account_import(False, False)
            input_login_info(aim_url)
        except Exception as e:
            print('logined error:{e}')

def input_login_info(aim_url):
    # login_url = "https://login.ti.com/as/authorization.oauth2?response_type=code&scope=openid%20email%20profile&client_id=DCIT_ALL_WWW-PROD&state=6EgvgF3zTyFQhyXOWnbLBODAshQ&redirect_uri=https%3A%2F%2Fwww.ti.com.cn%2Foidc%2Fredirect_uri%2F&nonce=1REaCAcnxx0ymCZum3eIcJYDP2T4zce24nzBtjbB_3E&response_mode=form_post"
    user_name = "zhouhaojian@kehua.com"
    user_password = "Aa147147"
    driver.find_element(by=By.ID, value='username').send_keys(user_name)
    WaitHelp.waitfor_account_import(False, False)
    driver.find_element(by=By.ID, value='nextbutton').click()
    time.sleep(4)
    while not driver.find_element(by=By.ID, value='password'):
        time.sleep(1.0)
    driver.find_element(by=By.ID, value='password').find_element(by=By.TAG_NAME, value='input').send_keys(user_password)
    WaitHelp.waitfor_account_import(False, False)
    driver.find_element(by=By.ID, value='loginbutton').click()
    WaitHelp.waitfor_account_import(False, False)
    if driver.current_url == "https://www.ti.com.cn/":
        driver.get(aim_url)
        WaitHelp.waitfor_account_import(True, False)


# 查询型号的库存信息，有则加到购物车
def use_stock(cate_index, cate_name):
    try:
        driver.find_element(by=By.CLASS_NAME, value="magic-box-input").find_element(by=By.TAG_NAME,
                                                                                    value='input').clear()
        driver.find_element(by=By.CLASS_NAME, value="magic-box-input").find_element(by=By.TAG_NAME,
                                                                                    value='input').send_keys(cate_name)
        try:
            driver.find_element(by=By.CLASS_NAME, value="coveo-search-button").click()
        except Exception as e:
            LogHelper.write_log(log_file_name=log_file, content=f'{cate_name} can not get search button exception: {e}')
        WaitHelp.waitfor_account_import(False, False)
        WaitHelp.waitfor_account_import(False, False)
        loop_times = 0
        while loop_times <= 5 and driver.current_url.find(f'searchTerm={cate_name}') == -1:
            print('searchTerm=', cate_name)
            print("url is :", driver.current_url)
            loop_times += 1
            WaitHelp.waitfor_account_import(False, False)
        snap_kind = use_snapshot_kind()
        if snap_kind == 1:
            analyth_opn_snapshot(cate_name=cate_name)
        elif snap_kind == 2:
            analyth_tool_snapshot(cate_name=cate_name)
        elif snap_kind == 3:
            LogHelper.write_log(log_file_name=log_file, content=f'{cate_name } snap_kind == 3')
    except Exception as e:
        LogHelper.write_log(log_file_name=log_file, content=f'{cate_name} can not find search input box exception: {e}')


# 解析 ti_opn_snapshot, 获取 limit 总数
def analyth_opn_snapshot(cate_name) -> int:
    try:
        search_snap = driver.find_element(by=By.CSS_SELECTOR, value='div.search-query-snapshots')
        ti_opn_snapshot = search_snap.find_element(By.CSS_SELECTOR, value='ti-opn-snapshot')
        snap_shadow = ti_opn_snapshot.shadow_root
        match_details = snap_shadow.find_element(By.CSS_SELECTOR, value='div.ti-opn-snapshot-opn-match-details')
        opn_details = match_details.find_elements(By.CSS_SELECTOR, value='ti-opn-details')
        for one_row in opn_details:
            row_shadow = one_row.shadow_root
            opn_detail_str = one_row.get_attribute('opn-detail-data')
            detail_dic = json.loads(opn_detail_str)
            opn_detail_data = detail_dic['opnParameters']['distributorsList']
            limit_des = row_shadow.find_element(By.CSS_SELECTOR, value='ti-tooltip-trigger').text
            if re.sub("\D", "", limit_des):
                limit_num = int(re.sub("\D", "", limit_des))
                add_cart_opn_snapshot(cate_name=cate_name, limit_stock_number=limit_num, one_row=one_row, detail_data= str(opn_detail_data))
                record_stock_info(stock_number=limit_num, cate_name=cate_name, detail_data=str(opn_detail_data))
    except Exception as e:
        LogHelper.write_log(log_file_name=log_file, content=f'{cate_name} analyze_opn_snapshot Exception: {e}')


# 解析 ti_opn_snapshot, 加入购物车，关闭弹框
def add_cart_opn_snapshot(cate_name, limit_stock_number, one_row, detail_data):
    if limit_stock_number > 100:
        EmailHelper.mail_TI(cate_name, limit_stock_number, detail_data)
        try:
            row_shadow = one_row.shadow_root
            ti_add_to_card_shadow = row_shadow.find_element(by=By.CSS_SELECTOR,
                                                                         value='ti-add-to-cart').shadow_root
            ti_input_shadow = ti_add_to_card_shadow.find_element(by=By.CSS_SELECTOR, value='ti-input').shadow_root
            ti_input_shadow.find_element(by=By.CSS_SELECTOR, value='input').clear()
            ti_input_shadow.find_element(by=By.CSS_SELECTOR, value='input').send_keys(str(limit_stock_number))
            WaitHelp.waitfor_account_import(False, False)
            ti_add_to_card_shadow.find_element(by=By.CSS_SELECTOR, value='ti-button').click()
            WaitHelp.waitfor_account_import(False, False)
            close_add_cart()
        except Exception as e:
            LogHelper.write_log(log_file_name=log_file, content=f'{cate_name} add card error exception: {e}')
    else:
        print('not enough stock')


# 解析tool_snapshot 获取库存数量
def analyth_tool_snapshot(cate_name):
    try:
        ti_tool_snapshot_shadow = driver.find_element(by=By.TAG_NAME, value='ti-tool-snapshot').shadow_root
        limit_arr = ti_tool_snapshot_shadow.find_elements(by=By.CLASS_NAME, value='ti-tool-snapshot-store-info')
        if len(limit_arr) > 0:
            limit_ele = ti_tool_snapshot_shadow.find_elements(by=By.CLASS_NAME, value='ti-tool-snapshot-store-info')[-1]
            limit_des = limit_ele.text
            if re.sub("\D", "", limit_des):
                limit_number = int(re.sub("\D", "", limit_des))
                add_cart_tool_snapshot(cate_name=cate_name, limit_stock_number=limit_number)
                # tool_snapshot 没有detail_data这段数据
                record_stock_info(stock_number=limit_number, cate_name=cate_name, detail_data='--')
    except Exception as e:
        LogHelper.write_log(log_file_name=log_file, content=f'{cate_name} get_stock_number exception: {e}')


# ti_tool_snapshot, 加购物车
def add_cart_tool_snapshot(cate_name, limit_stock_number):
    if limit_stock_number > 100:
        EmailHelper.mail_TI(cate_name, limit_stock_number, '--')
        try:
            ti_tool_snapshot_shadow = driver.find_element(by=By.TAG_NAME, value='ti-tool-snapshot').shadow_root
            ti_add_to_card_shadow = ti_tool_snapshot_shadow.find_element(by=By.CSS_SELECTOR,
                                                                         value='ti-add-to-cart').shadow_root
            ti_input_shadow = ti_add_to_card_shadow.find_element(by=By.CSS_SELECTOR, value='ti-input').shadow_root
            ti_input_shadow.find_element(by=By.CSS_SELECTOR, value='input').clear()
            ti_input_shadow.find_element(by=By.CSS_SELECTOR, value='input').send_keys(str(limit_stock_number))
            WaitHelp.waitfor_account_import(False, False)
            ti_add_to_card_shadow.find_element(by=By.CSS_SELECTOR, value='ti-button').click()
            WaitHelp.waitfor_account_import(False, False)
            close_add_cart()
        except Exception as e:
            LogHelper.write_log(log_file_name=log_file, content=f'{cate_name} add card error exception: {e}')
    else:
        print('not enough stock')


# close add cart alert
def close_add_cart():
    #  关闭弹框
    try:
        ti_dialog_shadow = driver.find_element(by=By.TAG_NAME, value='ti-dialog').shadow_root
        ti_svg_icon_shadow = ti_dialog_shadow.find_element(by=By.CSS_SELECTOR, value='ti-svg-icon').shadow_root
        ti_svg_icon_shadow.find_element(by=By.CLASS_NAME, value='ti-svg-icon-bg').click()
        WaitHelp.waitfor_account_import(False, False)
    except Exception as e:
        print(f"close alert view exception:{e}")


# 记录库存查询结果
def record_stock_info(stock_number, cate_name, detail_data):
    save_file = '//TI/TI_cate_add.xlsx'
    IC_Stock_excel_write.add_arr_to_sheet(file_name=save_file, sheet_name='stock', dim_arr=[[cate_name,
                                                                                             stock_number,
                                                                                             time.strftime("%Y-%m-%d %H:%M:%S", time.localtime()),
                                                                                             detail_data]])


# 获取garde == A 的所有[[cate_name, manufacture]]
def get_A_cateAndManu() -> list:
    result = []
    gradeA_file = '//TgardeA.xlsx'
    wb = load_workbook(filename=gradeA_file)
    ws = wb['before0910']
    # 根据单元格名称获取单元格对象
    for i in range(ws.min_row, ws.max_row + 1):
        cate_name = ws.cell(i, 1).value
        manufacture = ws.cell(i, 2).value
        if manufacture == 'Texas Instruments' or manufacture == 'TI':
            result.append(cate_name)
    return result


def get_json_cate(file_name: object, sheet_name: object, col_index: int) -> list:
    # 获取工作簿对象
    wb = load_workbook(filename=file_name)
    # 获取sheet
    ws = wb[sheet_name]
    # 根据单元格名称获取单元格对象
    result = []
    for i in range(ws.min_row, ws.max_row + 1):
        cate = ws.cell(i, col_index).value
        manufacture = ws.cell(i, 2).value
        if not cate == "--":
            if manufacture == 'Texas Instruments' or manufacture == 'TI':
                result.append(cate)
    return result


# 获取grade 为A，且 manu 为TI的cate
def get_buy_cates():
    gradeA_cates = get_A_cateAndManu()
    json_cates = get_json_cate(file_name=cate_source_file, sheet_name='all', col_index=1)
    all_cates = list(set(gradeA_cates).union(set(json_cates)))
    return all_cates


def main():
    cate_ids = get_buy_cates()
    black_cates = IC_stock_excel_read.get_cate_name_arr(file_name=cate_source_file, sheet_name='blacklist', col_index=1)
    cate_ids = list(set(cate_ids).difference(set(black_cates)))
    # cate_ids = ['TLC5947RHBR', 'TCAN1043DEVM', 'AMC1306M05QDWVRQ1']
    # 不停的遍历是否有库存
    print(cate_ids)
    while True:
        for (cate_index, cate_name) in enumerate(cate_ids):
            print(f"index is: {cate_index}, cate_name is: {cate_name}")

            if cate_index > 0 and cate_index%15 == 0:
                time.sleep(480)
            else:
                WaitHelp.waitfor_account_import(True, False)
            use_stock(cate_index, cate_name)


if __name__ == "__main__":
    driver.get(default_url)
    WaitHelp.waitfor_account_import(False, False)
    login_action(default_url)
    main()
